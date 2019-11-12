import openpyxl as xls
import pandas as pd
import time
from functools import lru_cache
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

class ExcelWorkBookWriter:
  def __del__(self):
    self.book.close()

  def __init__(self, path_to_workbook, read_only = True):

    self.file_name = path_to_workbook
    self.read_only = read_only
    try:
      self.book = xls.load_workbook(path_to_workbook, read_only = read_only, data_only = True)

    except FileNotFoundError:
      logging.error('Workbook %s not found', path_to_workbook)
      self.book = None
    except:
      logging.error('Problem with loading xlsx file %s', path_to_workbook)
      self.book = None

  
  def get_working_sheet(self, sheet_name):
    try:
        return self.book[sheet_name]
    except:
        logging.error('Sheet %s not found', sheet_name)
        return None

  @lru_cache(maxsize=None)
  def get_header(self, sheet_name, header_row):
    ws = self.get_working_sheet(sheet_name)
    if ws is not None:
        row = ws[header_row]
        header = [cell.value for cell in row]
        header  = map(lambda x: x.upper() if x is not None else "_", header)
        return list(header)
    else:
        return None
 
  def update_workbook(self, sheet_name, df, header, start_row, column_mapping_dict):
      logging.info('Starting to update %s', sheet_name)

      working_sheet = self.book[sheet_name]
      
      if df.shape[0] > 0 :
        for i, r in df.iterrows():
            for k, v in column_mapping_dict.items():
                try:
                    k_in_excel = header.index(k.upper())+1
                except ValueError:
                    logging.error("Column {} not found".format(k))
                else:
                    if v in r.index:
                        value = self.get_bom_translated_value(r[v], v)
                        working_sheet.cell(start_row, k_in_excel).value = value 
            
            # bolt length-> ino DN2 columns
            if r['PART_TYPE'] == 'BOLT':
                k_in_excel = header.index('DN1')+1    
                working_sheet.cell(start_row, k_in_excel).value = r['BOLT-DIA']
                k_in_excel = header.index('DN2')+1    
                working_sheet.cell(start_row, k_in_excel).value = r['BOLT-LENGTH']
                k_in_excel = header.index('DN2 EINHEIT')+1
                working_sheet.cell(start_row, k_in_excel).value = r['UNITS-BOLT-LENGTH'].lower()

            start_row += 1
      
      return
    
  def get_bom_translated_value(self, value, column_name):
        if value == "INCH": return "in"
        elif value == "MM": return "mm"
        elif str(value).startswith('0.') and column_name in ['DN1', 'DN2']: return str(value)[1:]
        elif str(value).endswith('.0000'): return str(value)[:-5]
        elif str(value).endswith('000') and column_name in ['DN1', 'DN2']: return str(value)[:-3]
        elif str(value).endswith('00') and column_name in ['DN1', 'DN2']: return str(value)[:-2]
        else:
            return value
